import pandas as pd
import time
import pickle

from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime
from win32com.client import Dispatch

# Index(['detail_Project', 'detail_Project_Long_Name', 'detail_Phase',
#        'detail_Task', 'detail_Task_Name', 'detail_Organization_Name',
#        'detail_Principal_Name', 'detail_Project_Manager_Name',
#        'detail_Project_Name', 'detail_Primary_Client',
#        'detail_Bill_Client_Name', 'detail_Biller_Name', 'detail_Status'],
#       dtype='object')

# Script run Daily through Task Scheduler


class TedDaily:
    def __init__(self):

        # Project codes: add if necessary
        # Boolean: 1-Reg 0-Promotional
        # Dict/mapping for create_daily
        self.WORK_ORDER = (
            ('1403.000.001', True), ('1403.000.999', False),
            ('1404.000.001', True), ('1404.000.999', False),
            ('1405.000.001', True), ('1405.000.999', False),
            ('1411.000.001', True), ('1411.000.999', False),
            ('1412.000.001', True), ('1412.000.999', False),
            ('1413.000.001', True), ('1413.000.999', False),
            ('1421.000.001', True), ('1421.000.999', False),
        )

        self.DIR_DICT = dict()  # Where pickled data is stored. mostly for path
        self.unpickle()  # Func to unpickle data
        self.TO_FPATH = f'{self.DIR_DICT["FROM_HOST"]}\_FROMOUTLOOK\_TED'  # Data from pickle
        self.TED_FPATH = f'{self.DIR_DICT["BOX_MAIN"]}\_REPORTS'  # Data from pickle
        self.TIME_STAMP = datetime.now().strftime('%Y%m%d')

        # ITD variables
        self.itd_col_names = [
            'PROJNUM', 'INVAMT', 'INVNUM',
            'PO', 'MPER', 'YPER',
        ]

        # self.filter_list = {
        #         '1403.000.001': ['5660067010', 'Transmission TED Reliability Program OE Services', '5000000', ],
        #         '1404.000.001': ['5660067009', 'Transmission TED CMP Program OE Services', '5000000', ],
        #         '1405.000.001': ['5660067008', 'Transmission TED Fiber Program OE Services', '5000000', ],
        #         '1411.000.001': ['5660067010', 'Distribution TED Reliability Program OE Services', '5000000', ],
        #         '1412.000.001': ['5660067009', 'Distribution TED CMP Program OE Services', '5000000', ],
        #         '1413.000.001': ['5660067008', 'Distribution TED Fiber Program OE Services', '5000000', ],
        #         }

        self.itd_data = pd.read_csv(fr'{self.DIR_DICT["MAIN_TEMP"]}\ITD_VS.csv',
                                    header=None,
                                    names=self.itd_col_names,
                                    )

        self.filtered_itd = pd.DataFrame()
        self.po_list = list()
        self.xl_file = Workbook()

        # PM: to company PM
        # DISTR: to non-PM project members
        # to, cc, filename,

        self.outlook_wo = {
        'PM': [
            'fulano.olivera@company.com; fulano.olivera@company.com',
            '',
            'TED_PROJ_STRUCTURE.xlsx',
        ],

        'DISTR': [
            'john.doe@company.com; john.doe@company.com',
            'TED_PROJ_STRUCTURE_DISTR.xlsx',
        ],
            }

        self.output_save_dir = self.DIR_DICT['MAIN_TEMP']

    def unpickle(self):

        # Cheating, using the error as a way to get the other path so script can be executed through BAT
        with open(r'C:\Users\V Song\Desktop\_PYTHONSCRIPTS\_picklenv\data\envs.pickle', 'rb') as pfile:
            self.DIR_DICT = pickle.load(pfile)

        return

    def create_daily(self):
        df = pd.DataFrame()
        source_df = pd.read_csv(f"{self.DIR_DICT['FROM_HOST']}\_FROMOUTLOOK\PROJECT_LIST_EXPORT.csv")

        # conditions/filters for the TED projects
        for projname in self.WORK_ORDER:
            if projname[1]:
                filter1 = source_df["detail_Project"].str.contains(projname[0])
                filter2 = source_df["detail_Task"].str.contains(" ")
                filter3 = source_df["detail_Project_Long_Name"].str.contains("SDGE") == False
                filter4 = source_df["detail_Project_Long_Name"].str.contains("SUBCONTRACTOR") == False
                filter5 = source_df["detail_Status"].str.contains("Active")

                # Apply all filters into raw csv
                temp_df = source_df[filter1 & filter2 & filter3 & filter4 & filter5]

                # print(temp_df)
                df = df.append(temp_df, ignore_index=True)

                del temp_df
                del filter1
                del filter2
                del filter3
                del filter4
                del filter5

        # print(df)
        # df = df.drop(columns=[
        #                     'detail_Task', 'detail_Organization_Name',
        #                     'detail_Principal_Name', 'detail_Project_Manager_Name', 'detail_Project_Name',
        #                     'detail_Primary_Client', 'detail_Bill_Client_Name', 'detail_Biller_Name',
        #             ])

        df = df.drop(columns=[
            'detail_Task', 'detail_Organization_Name',
            'detail_Project_Manager_Name', 'detail_Project_Name',
            'detail_Primary_Client', 'detail_Biller_Name',
        ])

        # to re order columns
        df = df[['detail_Project', 'detail_Phase', 'detail_Project_Long_Name', 'detail_Task_Name', 'detail_Status', ]]

        df = df.rename(columns={
            'detail_Project': 'PROJNUM',
            'detail_Phase': 'TASK',
            'detail_Project_Long_Name': 'DESC',
            'detail_Task_Name': 'SHRTND',
            'detail_Status': 'STATUS',
        })

        self.filtered_itd = df.copy()

        df.to_csv(
            fr'{self.output_save_dir}\TED_PROJ_LIST.csv',
            index=False
        )

        return

    def prep_excel_ted(self):

        # unpacking self.WORK_ORDER,
        # only non .999 projects for TED
        ted_projs = [x[0] for x in self.WORK_ORDER if x[1]]  # extract all true in self.workorder
        # ted_projs['INVAMT'] = ted_projs['INVAMT'].map(lambda x: '{:,.2f}'.format(x))
        # filter data based on deltek ted_projs codes 1403 - 1405/ 1411 - 1413
        ted_itd = self.itd_data[self.itd_data['PROJNUM'].str.contains('|'.join(ted_projs))]  # filtered using re & join

        # ted_itd = ted_itd[ted_itd['INVAMT'].map(lambda x: '{:,.2f}'.format(x))]


        # Opening dataframes and existing workbook
        wbs_df = pd.read_csv(fr'{self.output_save_dir}\TED_PROJ_LIST.csv')
        wb = load_workbook(fr'{self.output_save_dir}\TED_PROJ_STRUCTURE.xlsx')

        # Workorder for dict loop for the TASK and ITD sheet in excel file DISTR
        # used in the following for loop
        # 'SHEETNAME': [PROJLIST, NUMCOLS]

        print(ted_itd)

        wo_dict = {
            'TASK': [wbs_df, 4],
            'ITD': [ted_itd, 6],
        }

        for k, v in wo_dict.items():
            # Task sheet clean and update
            ws = wb[k]
            ws.delete_cols(1, v[1])

            rows = dataframe_to_rows(v[0], index=False, header=True)
            for i, x in enumerate(rows):
                for c in range(0, v[1]):
                    ws.cell(row=i+1, column=c+1).value = x[c]

        wb.save(fr'{self.output_save_dir}\TED_PROJ_STRUCTURE.xlsx')

        sheet_to_del = wb.get_sheet_by_name('NOTES')
        wb.remove_sheet(sheet_to_del)

        sheet_to_del = wb.get_sheet_by_name('ITD')
        wb.remove_sheet(sheet_to_del)

        wb.save(fr'{self.output_save_dir}\TED_PROJ_STRUCTURE_DISTR.xlsx')

        del wb

    def send_excel(self):

        olapp = Dispatch('Outlook.Application')

        for k, v in self.outlook_wo.items():

            olmail = olapp.CreateItem(0)
            olmail.To = v[0]
            olmail.Cc = v[1]
            olmail.Subject = f'DAILY_TED_PROJECT_STRUCTURE_{self.TIME_STAMP}_{k}'
            olmail.Attachments.Add(fr'{self.output_save_dir}\{v[2]}')

            olmail.HTMLbody = f'Hi -- <br><br>' \
                              f'Please see attached for the daily TED Project Structure Report as for {self.TIME_STAMP}<br><br>' \
                              f'Let me know if you find any discrepancies/duplicates' \
                              f'<br><br>' \
                              f'Thanks <br><br> -v'

            olmail.display(True)


def main():
    td = TedDaily()
    td.create_daily()
    time.sleep(0.5)
    td.prep_excel_ted()
    time.sleep(0.5)
    td.send_excel()


if __name__ == '__main__':
    main()
